from __future__ import annotations

import copy
import math
import re
import tempfile
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel, to_excel


BASE_DIR = Path(__file__).resolve().parent
MAIN_SHEET = "MAINCPF CPH DEP SEP END RBC CSR"
INPUT_START_COL = 1  # A
INPUT_END_COL = 56  # BD
CALC_START_COL = 58  # BF
CALC_END_COL = 91  # CM
OUTPUT_START_COL = 93  # CO
OUTPUT_END_COL = 101  # CW
HEADER_ROWS = 2
DATA_START_ROW = 3
BLANK_STREAK_STOP = 200
ACCRUED_INTEREST_DATE_SERIAL = float(to_excel(datetime(2026, 3, 31)))
ACCRUED_INTEREST_RATE = 0.0825
OUTPUT_HEADER_ROW1 = {
    93: "Protiviti",
    101: "x",
}
TEMPLATE_COLUMN_WIDTHS = {
    1: 12.42578125, 2: 10.0, 3: 14.42578125, 4: 12.7109375, 5: 13.42578125, 6: 12.0,
    7: 8.42578125, 8: 13.0, 9: 12.42578125, 10: 10.85546875, 11: 8.42578125,
    12: 9.28515625, 13: 13.7109375, 14: 13.42578125, 15: 13.7109375, 16: 9.28515625,
    17: 12.42578125, 18: 13.0, 19: 13.0, 20: 12.7109375, 21: 10.0, 22: 19.140625,
    23: 13.0, 24: 14.0, 25: 10.140625, 26: 11.42578125, 27: 13.28515625,
    28: 28.85546875, 29: 13.0, 30: 14.0, 31: 14.7109375, 32: 12.28515625,
    33: 14.85546875, 34: 13.0, 35: 16.0, 36: 11.0, 37: 13.0, 38: 11.28515625,
    39: 16.7109375, 40: 15.42578125, 41: 17.28515625, 42: 23.140625, 43: 27.0,
    44: 21.140625, 45: 28.42578125, 46: 15.42578125, 47: 17.42578125, 48: 18.42578125,
    49: 15.42578125, 50: 35.85546875, 51: 22.7109375, 52: 38.42578125,
    53: 17.28515625, 54: 17.140625, 55: 13.0, 56: 13.0, 93: 14.140625, 94: 22.42578125,
    95: 13.0, 96: 13.0, 97: 13.0, 98: 13.0, 99: 13.0, 100: 13.0, 101: 13.0,
}
OUTPUT_HEADER_ROW2 = {
    93: "LN_LA \nUpdated",
    94: "Guaranteed Addition",
    95: "GMA",
    96: "Overall Revisionary Bonus",
    97: "Special Bonus",
    98: "Terminal Bonus",
    99: "Maturity Benefit",
    100: "Condition of 100.1% , additional bonus for CSR and END",
    101: "Total Maturity Benefit (inclusive of Rev. Bonus wherever applicable)",
}
CPF_CPH_OPTIONS = {
    "SS": "(100%*BSA)+\nGA+GMA",
    "PG": "(50%*BSA)\n+GA",
    "PD": "(25%*BSA)\n+GA",
    "CS": "(20%*BSA)+\nGA",
}
CSR_HIGH_SA_TABLE = [
    (0.0, {10: 0.0, 13: 0.0, 16: 0.0, 19: 0.0, 22: 0.0}),
    (250000.0, {10: 0.04, 13: 0.06, 16: 0.08, 19: 0.10, 22: 0.12}),
    (500000.0, {10: 0.08, 13: 0.12, 16: 0.16, 19: 0.20, 22: 0.24}),
]
RBC_TERMINAL_BONUS_TABLE = {
    10: {10: 0.176, 11: 0.2222, 12: 0.2796, 13: 0.3484, 14: 0.4312, 15: 0.5355},
    11: {10: 0.0, 11: 0.2222, 12: 0.2796, 13: 0.3484, 14: 0.4312, 15: 0.5355},
    12: {10: 0.0, 11: 0.0, 12: 0.2796, 13: 0.3484, 14: 0.4312, 15: 0.5355},
}
DEP_BONUS_RATES = {
    1: 0.0, 2001: 0.0, 2002: 0.0, 2003: 0.0, 2004: 0.0, 2005: 0.0275, 2006: 0.03,
    2007: 0.0275, 2008: 0.025, 2009: 0.02, 2010: 0.02, 2011: 0.02, 2012: 0.018,
    2013: 0.018, 2014: 0.018, 2015: 0.018, 2016: 0.018, 2017: 0.018, 2018: 0.018,
    2019: 0.018, 2020: 0.018, 2021: 0.018, 2022: 0.018, 2023: 0.018, 2024: 0.018,
    2025: 0.018, 2026: 0.018,
}
SEP_BONUS_RATES = {
    2001: 0.0, 2002: 0.03, 2003: 0.035, 2004: 0.03, 2005: 0.0275, 2006: 0.03,
    2007: 0.03, 2008: 0.03, 2009: 0.026, 2010: 0.026, 2011: 0.026, 2012: 0.023,
    2013: 0.023, 2014: 0.023, 2015: 0.021, 2016: 0.04, 2017: 0.04, 2018: 0.04,
    2019: 0.04, 2020: 0.04, 2021: 0.04, 2022: 0.04, 2023: 0.04, 2024: 0.036,
    2025: 0.034, 2026: 0.034,
}
END_BONUS_RATES = {
    1: 0.0, 2001: 0.0, 2002: 0.0, 2003: 0.0, 2004: 0.0, 2005: 0.0, 2006: 0.0,
    2007: 0.0, 2008: 0.0, 2009: 0.0, 2010: 0.0, 2011: 0.0, 2012: 0.0, 2013: 0.018,
    2014: 0.0125, 2015: 0.018, 2016: 0.02, 2017: 0.02, 2018: 0.02, 2019: 0.02,
    2020: 0.02, 2021: 0.02, 2022: 0.02, 2023: 0.02, 2024: 0.02, 2025: 0.02,
    2026: 0.02,
}
RBC_BONUS_RATES = {
    2001: 0.0, 2002: 0.0, 2003: 0.0, 2004: 0.0, 2005: 0.0, 2006: 0.0, 2007: 0.0,
    2008: 0.0, 2009: 0.0, 2010: 0.0, 2011: 0.0, 2012: 0.0, 2013: 0.0, 2014: 0.0,
    2015: 0.04, 2016: 0.055, 2017: 0.055, 2018: 0.055, 2019: 0.055, 2020: 0.055,
    2021: 0.055, 2022: 0.055, 2023: 0.055, 2024: 0.055, 2025: 0.055, 2026: 0.055,
}
CSR_BONUS_RATES = {
    2001: 0.0, 2002: 0.0, 2003: 0.0, 2004: 0.0, 2005: 0.0, 2006: 0.0, 2007: 0.0,
    2008: 0.0, 2009: 0.0, 2010: 0.0, 2011: 0.0, 2012: 0.0, 2013: 0.0, 2014: 0.023,
    2015: 0.021, 2016: 0.025, 2017: 0.025, 2018: 0.025, 2019: 0.03, 2020: 0.03,
    2021: 0.03, 2022: 0.03, 2023: 0.03, 2024: 0.03, 2025: 0.03, 2026: 0.03,
}
THIN_SIDE = Side(style="thin", color="FF000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
STYLE_PRESETS = {
    "plain": {
        "font": Font(name="Aptos Narrow", size=11),
        "fill": PatternFill(fill_type=None),
        "border": Border(),
        "alignment": Alignment(),
        "number_format": "General",
    },
    "plain_bold": {
        "font": Font(name="Aptos Narrow", size=11, bold=True),
        "fill": PatternFill(fill_type=None),
        "border": Border(),
        "alignment": Alignment(),
        "number_format": "General",
    },
    "plain_date": {
        "font": Font(name="Aptos Narrow", size=11),
        "fill": PatternFill(fill_type=None),
        "border": Border(),
        "alignment": Alignment(),
        "number_format": "d-mmm-yy",
    },
    "plain_amount": {
        "font": Font(name="Aptos Narrow", size=11),
        "fill": PatternFill(fill_type=None),
        "border": Border(),
        "alignment": Alignment(),
        "number_format": '_ * #,##0.00_ ;_ * \\-#,##0.00_ ;_ * "-"??_ ;_ @_ ',
    },
    "hdr": {
        "font": Font(name="Aptos Narrow", size=11, bold=True),
        "fill": PatternFill(fill_type=None),
        "border": THIN_BORDER,
        "alignment": Alignment(vertical="center"),
        "number_format": "General",
    },
    "hdr_date": {
        "font": Font(name="Aptos Narrow", size=11, bold=True),
        "fill": PatternFill(fill_type=None),
        "border": THIN_BORDER,
        "alignment": Alignment(vertical="center"),
        "number_format": "d-mmm-yy",
    },
    "hdr_yellow": {
        "font": Font(name="Aptos Narrow", size=11, bold=True),
        "fill": PatternFill(fill_type="solid", fgColor="FFFFFF00"),
        "border": THIN_BORDER,
        "alignment": Alignment(vertical="center"),
        "number_format": "General",
    },
    "hdr_amount": {
        "font": Font(name="Aptos Narrow", size=11, bold=True),
        "fill": PatternFill(fill_type=None),
        "border": THIN_BORDER,
        "alignment": Alignment(vertical="center"),
        "number_format": '_ * #,##0.00_ ;_ * \\-#,##0.00_ ;_ * "-"??_ ;_ @_ ',
    },
    "out_hdr_light": {
        "font": Font(name="Aptos Narrow", size=11, bold=True),
        "fill": PatternFill(fill_type="solid", fgColor="FFD9E2F3"),
        "border": THIN_BORDER,
        "alignment": Alignment(vertical="center", wrap_text=True),
        "number_format": "General",
    },
    "out_hdr_teal": {
        "font": Font(name="Aptos Narrow", size=11, bold=True),
        "fill": PatternFill(fill_type="solid", fgColor="FF00B050"),
        "border": THIN_BORDER,
        "alignment": Alignment(vertical="center", wrap_text=True),
        "number_format": "General",
    },
    "data": {
        "font": Font(name="Aptos Narrow", size=11),
        "fill": PatternFill(fill_type=None),
        "border": THIN_BORDER,
        "alignment": Alignment(),
        "number_format": "General",
    },
    "data_green": {
        "font": Font(name="Aptos Narrow", size=11),
        "fill": PatternFill(fill_type="solid", fgColor="FF92D050"),
        "border": THIN_BORDER,
        "alignment": Alignment(),
        "number_format": "General",
    },
    "data_date": {
        "font": Font(name="Aptos Narrow", size=11),
        "fill": PatternFill(fill_type=None),
        "border": THIN_BORDER,
        "alignment": Alignment(),
        "number_format": "d-mmm-yy",
    },
    "data_amount": {
        "font": Font(name="Aptos Narrow", size=11),
        "fill": PatternFill(fill_type=None),
        "border": THIN_BORDER,
        "alignment": Alignment(),
        "number_format": '_ * #,##0.00_ ;_ * \\-#,##0.00_ ;_ * "-"??_ ;_ @_ ',
    },
}
HEADER_STYLE_BY_SOURCE_COL = {
    1: "hdr", 2: "hdr", 3: "hdr", 4: "hdr", 5: "hdr", 6: "hdr", 7: "hdr", 8: "hdr",
    9: "hdr", 10: "hdr", 11: "hdr", 12: "hdr", 13: "hdr_yellow", 14: "hdr_yellow",
    15: "hdr_yellow", 16: "hdr", 17: "hdr", 18: "hdr", 19: "hdr", 20: "hdr", 21: "hdr",
    22: "hdr_date", 23: "hdr_date", 24: "hdr", 25: "hdr", 26: "hdr", 27: "hdr",
    28: "hdr", 29: "hdr", 30: "hdr", 31: "hdr", 32: "hdr_date", 33: "hdr_date",
    34: "hdr_date", 35: "hdr_date", 36: "hdr_date", 37: "hdr_date", 38: "hdr",
    39: "hdr", 40: "hdr", 41: "hdr", 42: "hdr", 43: "hdr", 44: "hdr", 45: "hdr",
    46: "hdr_amount", 47: "hdr", 48: "hdr", 49: "hdr", 50: "hdr", 51: "hdr",
    52: "hdr", 53: "hdr", 54: "hdr", 55: "hdr", 56: "hdr", 93: "out_hdr_light",
    94: "out_hdr_teal", 95: "out_hdr_teal", 96: "out_hdr_teal", 97: "out_hdr_teal",
    98: "out_hdr_teal", 99: "out_hdr_teal", 100: "out_hdr_teal", 101: "out_hdr_teal",
}
DATA_STYLE_BY_SOURCE_COL = {
    1: "data_green", 22: "data_date", 32: "data_date", 33: "data_date", 34: "data_date",
    35: "data_date", 36: "data_date", 46: "data_amount", 93: "data_amount",
    94: "data_amount", 95: "data_amount", 96: "data_amount", 97: "data_amount",
    98: "data_amount", 99: "data_amount", 100: "data_amount", 101: "data_amount",
}


class WorkbookProcessingError(RuntimeError):
    pass


@dataclass
class ProcessingResult:
    output_path: Path
    record_count: int


@dataclass
class WorkbookConfig:
    accrued_interest_date_serial: float
    accrued_interest_rate: float
    cpf_cph_options: dict[str, Any]
    csr_high_sa_table: list[tuple[float, dict[int, float]]]
    rbc_terminal_bonus_table: dict[int, dict[int, float]]
    dep_bonus_rates: dict[int, float]
    sep_bonus_rates: dict[int, float]
    end_bonus_rates: dict[int, float]
    rbc_bonus_rates: dict[int, float]
    csr_bonus_rates: dict[int, float]


@dataclass
class CalculationRow:
    values: dict[str, Any]

    def get(self, col: str) -> Any:
        return self.values.get(col)

    def set(self, col: str, value: Any) -> Any:
        self.values[col] = value
        return value


def process_uploaded_workbook(upload_path: Path, output_dir: Path) -> ProcessingResult:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{uuid.uuid4()}_output.xlsx"

    try:
        record_count = build_output_workbook(upload_path, output_path)
    except Exception as exc:
        raise WorkbookProcessingError(f"Pure Python workbook processing failed: {exc}") from exc

    return ProcessingResult(output_path=output_path, record_count=record_count)


def build_output_workbook(source_path: Path, output_path: Path) -> int:
    input_workbook = load_workbook(source_path, data_only=True, read_only=False)
    try:
        input_ws = input_workbook[input_workbook.sheetnames[0]]
        config = default_config()

        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = MAIN_SHEET

        # The uploaded workbook supplies only raw input A:BD. Layout, headers,
        # rates, and styling come from the local reference workbook.
        column_plan = list(range(INPUT_START_COL, INPUT_END_COL + 1)) + list(
            range(OUTPUT_START_COL, OUTPUT_END_COL + 1)
        )

        rows = list(iter_input_rows(input_ws))
        if not rows:
            raise WorkbookProcessingError("No rows found in the input workbook.")

        for row_num, row_values in rows[:HEADER_ROWS]:
            for output_col, source_col in enumerate(column_plan, start=1):
                value = (
                    row_values[source_col - 1]
                    if source_col <= INPUT_END_COL
                    else OUTPUT_HEADER_ROW1.get(source_col)
                    if row_num == 1
                    else OUTPUT_HEADER_ROW2.get(source_col)
                )
                output_ws.cell(row=row_num, column=output_col).value = value

        apply_fallback_headers(output_ws, column_plan)

        data_rows = 0
        for source_row_num, row_values in rows[HEADER_ROWS:]:
            input_values = {
                get_column_letter(col): row_values[col - 1]
                for col in range(INPUT_START_COL, INPUT_END_COL + 1)
            }
            calc = calculate_row(input_values, config)
            for output_col, source_col in enumerate(column_plan, start=1):
                value = (
                    row_values[source_col - 1]
                    if source_col <= INPUT_END_COL
                    else calc.get(get_column_letter(source_col))
                )
                output_ws.cell(row=source_row_num, column=output_col).value = value
            data_rows += 1

        apply_output_formatting(output_ws, column_plan, rows[-1][0])
        output_ws.freeze_panes = "A3"
        output_wb.save(output_path)
        return data_rows
    finally:
        input_workbook.close()


def iter_input_rows(source_ws):
    blank_streak = 0
    for row_num, row in enumerate(
        source_ws.iter_rows(min_row=1, max_col=INPUT_END_COL, values_only=True),
        start=1,
    ):
        row_values = list(row)
        if row_num <= HEADER_ROWS:
            yield row_num, row_values
            continue

        if any(not is_blank(value) for value in row_values[:INPUT_END_COL]):
            blank_streak = 0
            yield row_num, row_values
        else:
            blank_streak += 1
            if blank_streak >= BLANK_STREAK_STOP:
                break

def default_config() -> WorkbookConfig:
    return WorkbookConfig(
        accrued_interest_date_serial=ACCRUED_INTEREST_DATE_SERIAL,
        accrued_interest_rate=ACCRUED_INTEREST_RATE,
        cpf_cph_options=CPF_CPH_OPTIONS,
        csr_high_sa_table=CSR_HIGH_SA_TABLE,
        rbc_terminal_bonus_table=RBC_TERMINAL_BONUS_TABLE,
        dep_bonus_rates=DEP_BONUS_RATES,
        sep_bonus_rates=SEP_BONUS_RATES,
        end_bonus_rates=END_BONUS_RATES,
        rbc_bonus_rates=RBC_BONUS_RATES,
        csr_bonus_rates=CSR_BONUS_RATES,
    )


def is_blank(value: Any) -> bool:
    return value is None or value == ""


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def eq_text(left: Any, right: str) -> bool:
    return clean_text(left).casefold() == right.casefold()


def is_one_of(value: Any, *options: str) -> bool:
    return any(eq_text(value, option) for option in options)


def num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime):
        return float(to_excel(value))
    if isinstance(value, date):
        return float(to_excel(datetime(value.year, value.month, value.day)))
    raw = str(value).strip().replace("\u00a0", "")
    if raw == "":
        return 0.0
    if raw.startswith("(") and raw.endswith(")"):
        raw = f"-{raw[1:-1]}"
    raw = raw.replace(",", "")
    raw = raw.rstrip("%")
    return float(raw)


def safe_num(value: Any) -> float:
    try:
        return num(value)
    except Exception:
        return 0.0


def excel_sum(*values: Any) -> float:
    total = 0.0
    for value in values:
        if isinstance(value, (list, tuple)):
            total += excel_sum(*value)
        else:
            total += safe_num(value)
    return total


def excel_round(value: Any, digits: int = 0) -> float:
    quant = Decimal("1").scaleb(-digits)
    rounded = Decimal(str(num(value))).quantize(quant, rounding=ROUND_HALF_UP)
    result = float(rounded)
    return int(result) if digits == 0 else result


def rounddown(value: Any, digits: int = 0) -> float:
    factor = 10 ** digits
    scaled = num(value) * factor
    result = math.floor(scaled) / factor if scaled >= 0 else math.ceil(scaled) / factor
    return int(result) if digits == 0 else result


def excel_date(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        return from_excel(value)
    raw = clean_text(value)
    if not raw:
        return None
    raw = raw.replace(".", "/")
    raw = re.sub(r"\s+", " ", raw)
    for fmt in (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d/%m/%y",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return from_excel(float(raw))


def date_serial(value: Any) -> float:
    dt = excel_date(value)
    return 0.0 if dt is None else float(to_excel(dt))


def datedif_years(start: Any, end: Any) -> int:
    start_dt = excel_date(start)
    end_dt = excel_date(end)
    if start_dt is None or end_dt is None:
        return 0
    years = end_dt.year - start_dt.year
    if (end_dt.month, end_dt.day) < (start_dt.month, start_dt.day):
        years -= 1
    return years


def date_add_days(value: Any, days: int) -> float:
    return date_serial(value) + days


def fiscal_year_from_rcd(value: Any) -> int:
    dt = excel_date(value)
    if dt is None:
        return 0
    return dt.year if dt.month <= 3 else dt.year + 1


def paid_to_fiscal_year(value: Any) -> int | str:
    dt = excel_date(value)
    if dt is None:
        return ""
    return dt.year - 1 if dt.month <= 3 else dt.year


def get_csr_high_sa_rate(config: WorkbookConfig, original_sa: Any, policy_term: Any) -> float:
    term = int(safe_num(policy_term))
    selected = config.csr_high_sa_table[0][1].get(term, 0.0)
    sa = safe_num(original_sa)
    for threshold, rates in config.csr_high_sa_table:
        if sa >= threshold:
            selected = rates.get(term, selected)
    return selected


def get_rbc_terminal_bonus_rate(config: WorkbookConfig, policy_term: Any, policy_year: Any) -> float:
    term = int(safe_num(policy_term))
    year = int(safe_num(policy_year))
    return config.rbc_terminal_bonus_table.get(term, {}).get(year, 0.0)


def dep_bonus(row: CalculationRow, config: WorkbookConfig) -> float:
    # DEP / SEP / END use a compounded running balance model in their support sheets.
    return compounded_bonus(
        original_sa=safe_num(row.get("AT")),
        start_year=fiscal_year_from_rcd(row.get("AF")),
        end_year=paid_to_fiscal_year(row.get("AG")),
        rates=config.dep_bonus_rates,
    )


def sep_bonus(row: CalculationRow, config: WorkbookConfig) -> float:
    ptd_fy = paid_to_fiscal_year(row.get("AG"))
    if ptd_fy == "":
        return 0.0
    final_year = 2026 if int(ptd_fy) + 5 == 2026 else int(ptd_fy)
    return compounded_bonus(
        original_sa=safe_num(row.get("AT")),
        start_year=fiscal_year_from_rcd(row.get("AF")),
        end_year=final_year,
        rates=config.sep_bonus_rates,
    )


def end_bonus(row: CalculationRow, config: WorkbookConfig) -> float:
    return compounded_bonus(
        original_sa=safe_num(row.get("AT")),
        start_year=fiscal_year_from_rcd(row.get("AF")),
        end_year=paid_to_fiscal_year(row.get("AG")),
        rates=config.end_bonus_rates,
    )


def rbc_bonus(row: CalculationRow, config: WorkbookConfig) -> float:
    # RBC / CSR bonus tabs apply flat year-wise rates on original SA instead.
    status = row.get("B")
    cutoff_date = row.get("AH") if eq_text(status, "IF") else row.get("AG")
    return flat_bonus(
        original_sa=safe_num(row.get("AT")),
        start_year=fiscal_year_from_rcd(row.get("AF")),
        end_year=paid_to_fiscal_year(cutoff_date),
        rates=config.rbc_bonus_rates,
    )


def csr_bonus(row: CalculationRow, config: WorkbookConfig) -> float:
    return flat_bonus(
        original_sa=safe_num(row.get("AT")),
        start_year=fiscal_year_from_rcd(row.get("AF")),
        end_year=paid_to_fiscal_year(row.get("AG")),
        rates=config.csr_bonus_rates,
    )


def compounded_bonus(
    original_sa: float,
    start_year: int,
    end_year: int | str,
    rates: dict[int, float],
) -> float:
    if end_year == "" or start_year <= 0:
        return 0.0
    running_base = original_sa
    total = 0.0
    for year in sorted(rates):
        if start_year <= year <= int(end_year):
            bonus = running_base * rates[year]
            total += bonus
            running_base += bonus
    return total


def flat_bonus(
    original_sa: float,
    start_year: int,
    end_year: int | str,
    rates: dict[int, float],
) -> float:
    if end_year == "" or start_year <= 0:
        return 0.0
    total = 0.0
    for year in sorted(rates):
        if start_year <= year <= int(end_year):
            total += original_sa * rates[year]
    return total


def annualised_premium_for_csr_end(product: Any, frequency: Any, inst_premium_with_gst: Any) -> float:
    freq = int(safe_num(frequency))
    premium = safe_num(inst_premium_with_gst)
    if eq_text(product, "CSR"):
        return {
            1: premium,
            2: premium / 0.52,
            4: premium / 0.265,
            12: premium / 0.0834,
        }.get(freq, 0.0)
    if eq_text(product, "END"):
        return {
            1: premium,
            2: premium / 0.51,
            4: premium / 0.26,
            12: premium / 0.0834,
        }.get(freq, 0.0)
    return 0.0


def csr_sb_paid(sum_assured: Any, term: Any, status: Any, product: Any) -> float:
    if not eq_text(product, "CSR") or not eq_text(status, "IF"):
        return 0.0
    return {
        10: safe_num(sum_assured) * 0.25,
        13: safe_num(sum_assured) * 0.45,
        16: safe_num(sum_assured) * 0.70,
        19: safe_num(sum_assured) * 1.00,
        22: safe_num(sum_assured) * 1.35,
    }.get(int(safe_num(term)), 0.0)


def calculate_row(input_values: dict[str, Any], config: WorkbookConfig) -> CalculationRow:
    row = CalculationRow(dict(input_values))

    # BF:CM reproduces the workbook's backend calculation block. CO:CW is then built
    # from those intermediate values and written to the reduced output workbook.
    bf = row.set("BF", row.get("B"))
    bg = row.set("BG", datedif_years(row.get("AF"), row.get("AH")))
    bh = row.set("BH", safe_num(row.get("AD")) * safe_num(row.get("AE")))
    bi = row.set(
        "BI",
        excel_round((date_serial(row.get("AG")) - date_serial(row.get("AF"))) / 365 * safe_num(row.get("AE")), 0),
    )
    bj = row.set("BJ", safe_num(bi) / safe_num(bh) if safe_num(bh) else 0.0)
    bk = row.set("BK", annualised_premium_for_csr_end(row.get("Z"), row.get("AE"), row.get("BB")))

    bl = row.set(
        "BL",
        safe_num(row.get("AT"))
        if eq_text(bf, "IF")
        else safe_num(row.get("AT")) * safe_num(bi) / safe_num(bh)
        if is_one_of(bf, "PU", "PM") and safe_num(bh)
        else "",
    )

    ah_serial = date_serial(row.get("AH"))
    bm = row.set(
        "BM",
        0.0
        if ah_serial <= config.accrued_interest_date_serial
        else (safe_num(row.get("M")) + safe_num(row.get("N")))
        * ((1 + config.accrued_interest_rate) ** ((ah_serial - config.accrued_interest_date_serial) / 365) - 1),
    )
    bn = row.set("BN", safe_num(row.get("N")) + safe_num(bm))
    bo = row.set("BO", rounddown(datedif_years(row.get("AF"), row.get("AG")), 0))

    bp = "NA"
    if eq_text(row.get("B"), "DH"):
        grace_days = 15 if safe_num(row.get("AE")) == 12 else 30
        if safe_num(bo) < 3 and date_add_days(row.get("AG"), grace_days) < date_serial(row.get("AK")):
            bp = "Lapsed Death"
        elif date_add_days(row.get("AG"), grace_days) >= date_serial(row.get("AK")):
            bp = "Inforce Death"
        else:
            bp = "Paidup Death"
    row.set("BP", bp)

    bq = row.set("BQ", excel_round((date_serial(row.get("AG")) - date_serial(row.get("AF"))) / 365, 0))
    row.set("BR", config.cpf_cph_options.get(clean_text(row.get("AY")), 0) if is_one_of(row.get("Z"), "CPF", "CPH") else 0)

    bs = row.set(
        "BS",
        0.02 * safe_num(row.get("AT")) * safe_num(row.get("AC"))
        if is_one_of(row.get("Z"), "CPF", "CPH") and eq_text(bf, "IF")
        else 0.02 * safe_num(row.get("AT")) * safe_num(bq)
        if is_one_of(row.get("Z"), "CPF", "CPH") and eq_text(bf, "PU")
        else 0.0,
    )
    bt = row.set(
        "BT",
        0.02 * safe_num(row.get("AT")) * safe_num(row.get("AC"))
        if is_one_of(row.get("Z"), "CPF", "CPH") and eq_text(row.get("AY"), "SS") and eq_text(bf, "IF")
        else 0.0,
    )
    bu = row.set(
        "BU",
        0.2 * safe_num(bl)
        if is_one_of(row.get("Z"), "CPF", "CPH") and is_one_of(bf, "IF", "PU") and eq_text(row.get("AY"), "CS")
        else 1.0 * safe_num(bl)
        if is_one_of(row.get("Z"), "CPF", "CPH") and is_one_of(bf, "IF", "PU") and eq_text(row.get("AY"), "SS")
        else 0.25 * safe_num(bl)
        if is_one_of(row.get("Z"), "CPF", "CPH") and is_one_of(bf, "IF", "PU") and eq_text(row.get("AY"), "PD")
        else 0.5 * safe_num(bl)
        if is_one_of(row.get("Z"), "CPF", "CPH") and is_one_of(bf, "IF", "PU") and eq_text(row.get("AY"), "PG")
        else 0.0,
    )
    bv = row.set(
        "BV",
        excel_sum(bs, bt, bu) if is_one_of(row.get("Z"), "CPF", "CPH") and is_one_of(bf, "IF", "PU") and not is_blank(bl) else 0.0
    )

    bw = row.set(
        "BW",
        0.07 * safe_num(row.get("AT")) * safe_num(row.get("AD"))
        if eq_text(row.get("Z"), "RBC") and eq_text(bf, "IF")
        else 0.07 * safe_num(row.get("AT")) * safe_num(row.get("AD")) * safe_num(bj)
        if eq_text(row.get("Z"), "RBC") and eq_text(bf, "PU")
        else 0.0,
    )
    bx = row.set(
        "BX",
        get_rbc_terminal_bonus_rate(config, row.get("AC"), datedif_years(row.get("AF"), row.get("AH"))) * safe_num(bl)
        if eq_text(row.get("Z"), "RBC") and eq_text(bf, "IF")
        else 0.0,
    )
    by = row.set("BY", rbc_bonus(row, config) if eq_text(row.get("Z"), "RBC") else 0.0)
    bz = row.set(
        "BZ",
        safe_num(bl) if eq_text(row.get("Z"), "RBC") and is_one_of(bf, "IF", "PU") and not is_blank(bl) else 0.0
    )

    ca = row.set("CA", dep_bonus(row, config) if eq_text(row.get("Z"), "DEP") else 0.0)
    cb = row.set(
        "CB",
        safe_num(bl) if eq_text(row.get("Z"), "DEP") and is_one_of(bf, "IF", "PU") and not is_blank(bl) else 0.0
    )

    cc = "NA"
    if eq_text(row.get("Z"), "SEP"):
        if safe_num(bg) == safe_num(row.get("AC")):
            cc = "Rev. Bonus Due"
        elif safe_num(row.get("AD")) == safe_num(bg):
            cc = "Sum Assured Due"
    row.set("CC", cc)
    cd = row.set(
        "CD",
        sep_bonus(row, config)
        if eq_text(row.get("Z"), "SEP") and cc == "Rev. Bonus Due"
        else safe_num(bl)
        if eq_text(row.get("Z"), "SEP") and cc == "Sum Assured Due"
        else 0.0,
    )

    ce = row.set("CE", csr_bonus(row, config) if eq_text(row.get("Z"), "CSR") else 0.0)
    cf = row.set(
        "CF",
        get_csr_high_sa_rate(config, row.get("AT"), row.get("AC")) * safe_num(bl) if eq_text(row.get("Z"), "CSR") else 0.0
    )
    cg = row.set(
        "CG",
        safe_num(cf) + safe_num(bl) if eq_text(row.get("Z"), "CSR") and is_one_of(bf, "IF", "PU") and not is_blank(bl) else 0.0
    )
    ch = row.set("CH", csr_sb_paid(row.get("AT"), row.get("AC"), row.get("B"), row.get("Z")))
    ci = row.set(
        "CI",
        max((safe_num(bk) * safe_num(row.get("AD")) * 1.001) - safe_num(ch) - excel_sum(ce, cf, cg), 0.0)
        if eq_text(row.get("Z"), "CSR") and eq_text(bf, "IF")
        else 0.0,
    )

    cj = row.set("CJ", end_bonus(row, config) if eq_text(row.get("Z"), "END") else 0.0)
    ck = row.set(
        "CK",
        max((safe_num(bk) * safe_num(row.get("AD")) * 1.001) - excel_sum(cj, 0, safe_num(bl)), 0.0)
        if eq_text(row.get("Z"), "END") and eq_text(bf, "IF")
        else 0.0,
    )
    cl = row.set(
        "CL",
        safe_num(bl) if eq_text(row.get("Z"), "END") and is_one_of(bf, "IF", "PU") and not is_blank(bl) else 0.0
    )
    row.set(
        "CM",
        max(excel_sum(cj, ck, cl), -safe_num(row.get("C")) * 1.001)
        if eq_text(row.get("Z"), "END") and eq_text(bf, "IF")
        else excel_sum(cj, ck, cl) - safe_num(row.get("Q")) - excel_sum(row.get("M"), row.get("N"), row.get("O"))
        if eq_text(row.get("Z"), "END") and eq_text(bf, "PU")
        else 0.0,
    )

    row.set("CO", bn)
    row.set("CP", excel_sum(bs, bw))
    row.set("CQ", bt)
    row.set("CR", excel_sum(by, ca, cd, ce, cj))
    row.set("CS", 0.0)
    row.set("CT", bx)
    row.set(
        "CU",
        0.0 if eq_text(row.get("Z"), "SEP") and cc == "Rev. Bonus Due" else excel_sum(bu, bz, cb, cd, cg, cl)
    )
    row.set("CV", excel_sum(ci, ck))
    row.set("CW", excel_sum(row.get("CP"), row.get("CQ"), row.get("CR"), row.get("CS"), row.get("CT"), row.get("CU"), row.get("CV")) - safe_num(row.get("Q")) - excel_sum(row.get("M"), row.get("CO")))

    return row


def process_workbook_file_bytes(file_bytes: bytes, output_dir: Path) -> ProcessingResult:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_path = Path(temp_file.name)
        temp_file.write(file_bytes)

    try:
        return process_uploaded_workbook(temp_path, output_dir)
    finally:
        temp_path.unlink(missing_ok=True)


def apply_fallback_headers(output_ws, column_plan: list[int]) -> None:
    # Users may upload a workbook where CO:CW was cleared. Keep the final output
    # headers stable in the download even when those source header cells are blank.
    for output_col, source_col in enumerate(column_plan, start=1):
        if source_col in OUTPUT_HEADER_ROW2 and is_blank(output_ws.cell(2, output_col).value):
            output_ws.cell(2, output_col).value = OUTPUT_HEADER_ROW2[source_col]


def apply_output_formatting(output_ws, column_plan: list[int], last_row: int) -> None:
    # Apply the template layout from Python constants so runtime processing has no
    # workbook dependency while preserving the same visual output.
    for output_col, source_col in enumerate(column_plan, start=1):
        target_letter = get_column_letter(output_col)
        output_ws.column_dimensions[target_letter].width = TEMPLATE_COLUMN_WIDTHS.get(source_col, 13.0)

    output_ws.row_dimensions[2].height = 75.0

    for row_num in range(1, last_row + 1):
        for output_col, source_col in enumerate(column_plan, start=1):
            target_cell = output_ws.cell(row=row_num, column=output_col)
            if row_num == 1:
                style_name = row1_style_name(source_col)
            elif row_num == 2:
                style_name = HEADER_STYLE_BY_SOURCE_COL.get(source_col, "hdr")
            else:
                style_name = DATA_STYLE_BY_SOURCE_COL.get(source_col, "data")
            apply_style_preset(target_cell, style_name)

    output_ws.sheet_format.defaultRowHeight = 15.0
    output_ws.sheet_format.baseColWidth = 8
    output_ws.page_margins.left = 0.7
    output_ws.page_margins.right = 0.7
    output_ws.page_margins.top = 0.75
    output_ws.page_margins.bottom = 0.75
    output_ws.page_margins.header = 0.3
    output_ws.page_margins.footer = 0.3
    output_ws.page_setup.orientation = "portrait"
    output_ws.page_setup.paperSize = 9


def row1_style_name(source_col: int) -> str:
    if source_col == 93:
        return "plain_bold"
    if source_col == 34:
        return "plain_date"
    if source_col == 46:
        return "plain_amount"
    return "plain"


def apply_style_preset(target_cell, style_name: str) -> None:
    preset = STYLE_PRESETS[style_name]
    target_cell.font = copy.copy(preset["font"])
    target_cell.fill = copy.copy(preset["fill"])
    target_cell.border = copy.copy(preset["border"])
    target_cell.alignment = copy.copy(preset["alignment"])
    target_cell.protection = copy.copy(Protection(locked=True, hidden=False))
    target_cell.number_format = preset["number_format"]


def copy_cell_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format
